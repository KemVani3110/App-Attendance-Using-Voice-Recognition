import os
import pickle
import numpy as np
import sounddevice as sd
import librosa
from sklearn.svm import SVC
from sklearn.preprocessing import LabelEncoder
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def record_voice(duration=5, sample_rate=44100):
    print("Recording...")
    recording = sd.rec(int(duration * sample_rate), samplerate=sample_rate, channels=1)
    sd.wait()  # Wait until recording is finished
    return recording.flatten(), sample_rate  # Return 1D array of audio

def extract_features(file_path_or_recording, sample_rate=44100, is_recording=False):
    if is_recording:
        recording = file_path_or_recording
        if np.all(recording == 0) or np.max(np.abs(recording)) < 0.01:
            return None  # No voice detected or volume too low
        y = recording
        sr = sample_rate
    else:
        y, sr = librosa.load(file_path_or_recording, sr=sample_rate)
    mfccs = librosa.feature.mfcc(y=y, sr=sr, n_mfcc=60)
    return np.mean(mfccs.T, axis=0)

def train_model(data_dir):
    features = []
    labels = []
    for user_dir in os.listdir(data_dir):
        user_path = os.path.join(data_dir, user_dir)
        if os.path.isdir(user_path):
            for file_name in os.listdir(user_path):
                file_path = os.path.join(user_path, file_name)
                feature = extract_features(file_path)
                if feature is not None:
                    features.append(feature)
                    labels.append(user_dir)

    X = np.array(features)
    y = np.array(labels)
    
    label_encoder = LabelEncoder()
    y_encoded = label_encoder.fit_transform(y)

    model = SVC(probability=True)
    model.fit(X, y_encoded)

    with open('voice_recognition_model.pkl', 'wb') as f:
        pickle.dump({'model': model, 'label_encoder': label_encoder}, f)

    return model, label_encoder

def recognize_with_initial_files(initial_features_list, model, label_encoder):
    user_scores = {user: [] for user in label_encoder.classes_}

    for features in initial_features_list:
        if features is not None:
            probabilities = model.predict_proba([features])[0]
            for i, user in enumerate(label_encoder.classes_):
                user_scores[user].append(probabilities[i])

    average_scores = {user: np.mean(scores) for user, scores in user_scores.items() if scores}
    sorted_scores = sorted(average_scores.items(), key=lambda x: x[1], reverse=True)

    best_match, max_score = sorted_scores[0] if sorted_scores else (None, 0)
    return best_match, max_score, sorted_scores, user_scores

def save_to_excel(data, filename="attendance_records.xlsx"):
    df = pd.DataFrame(data)

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Timestamp", "User", "Status"])
    else:
        wb = load_workbook(filename)
        ws = wb.active

        existing_data = pd.DataFrame(ws.values)
        existing_data.columns = existing_data.iloc[0]
        existing_data = existing_data[1:]

        for entry in data:
            user = entry['User']
            timestamp = entry['Timestamp']
            status = entry['Status']

            if user in existing_data['User'].values:
                idx = existing_data.index[existing_data['User'] == user].tolist()[0]
                existing_data.at[idx, 'Timestamp'] = timestamp
                existing_data.at[idx, 'Status'] = status
            else:
                new_row = pd.DataFrame([[timestamp, user, status]], columns=['Timestamp', 'User', 'Status'])
                existing_data = pd.concat([existing_data, new_row], ignore_index=True)

        ws.delete_rows(1, ws.max_row)
        for r in dataframe_to_rows(existing_data, index=False, header=True):
            ws.append(r)

    column_widths = {'A': 20, 'B': 20, 'C': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    for cell in ws["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    wb.save(filename)

def compare_initial_files():
    try:
        model, label_encoder = train_model(voice_samples_directory)
        initial_features_list = [extract_features(os.path.join(initial_directory, f"audio_{i}.wav")) for i in range(1, 11)]
        best_match, max_score, sorted_scores, user_scores = recognize_with_initial_files(initial_features_list, model, label_encoder)

        if max_score > 0.5:
            result = f"The user with the highest match is: {best_match} with confidence {max_score * 100:.2f}%\n\n"
            result += "Confidence scores for the 10 initial files:\n"
            for i, score in enumerate(user_scores[best_match], start=1):
                result += f"File {i}: {score * 100:.2f}%\n"
            if all(score > 0.5 for score in user_scores[best_match]):
                result += "\nAll 10 files have a similarity rate above 50% for the best match user."
            else:
                result += "\nNot all 10 files have a similarity rate above 50% for the best match user."
            result += "\n\nOther matches with lower confidence:\n"
            for user, score in sorted_scores[1:]:
                result += f"User: {user}, Confidence: {score * 100:.2f}%\n"
            messagebox.showinfo("Detailed Result", result)
        else:
            messagebox.showinfo("Result", "Voice not recognized.")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def record_and_recognize():
    try:
        model, label_encoder = train_model(voice_samples_directory)
        
        while True:
            recording, sample_rate = record_voice()
            features = extract_features(recording, sample_rate, is_recording=True)
            if features is None:
                messagebox.showwarning("Warning", "No voice detected or volume too low.")
                continue

            probabilities = model.predict_proba([features])[0]
            max_prob_index = np.argmax(probabilities)
            max_prob = probabilities[max_prob_index]

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if max_prob > 0.5:
                recognized_user = label_encoder.inverse_transform([max_prob_index])[0]
                status = "Present"
                messagebox.showinfo("Result", f"Recognized User: {recognized_user} with confidence {max_prob * 100:.2f}%")
                recognition_data = [{"Timestamp": timestamp, "User": recognized_user, "Status": status}]
                save_to_excel(recognition_data)  # Save or update the recognized user data
            else:
                status = "Unknown"
                messagebox.showinfo("Result", "Voice not recognized.")
                recognition_data = [{"Timestamp": timestamp, "User": "Unknown", "Status": status}]
                save_to_excel(recognition_data)  # Save or update the unknown user data

            continue_recognition = messagebox.askyesno("Continue", "Do you want to continue recognizing voices?")
            if not continue_recognition:
                break

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

def exit_program():
    root.destroy()

# Directories
initial_directory = 'initial_audio_files_khoi'
voice_samples_directory = 'voice_samples'

# Create main window
root = tk.Tk()
root.title("Attendance System")

# Set window size and background color
root.geometry("600x400")
root.configure(bg="#f0f0f0")

# Create a frame for the buttons
frame = tk.Frame(root, bg="#f0f0f0")
frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

# Create and style buttons
btn_compare = tk.Button(frame, text="Compare initial audio files with user samples", command=compare_initial_files, font=("Arial", 14), bg="#4CAF50", fg="white", width=40, height=2)
btn_compare.pack(pady=10)

btn_record = tk.Button(frame, text="Record and recognize attendance", command=record_and_recognize, font=("Arial", 14), bg="#2196F3", fg="white", width=40, height=2)
btn_record.pack(pady=10)

btn_exit = tk.Button(frame, text="Exit", command=exit_program, font=("Arial", 14), bg="#f44336", fg="white", width=40, height=2)
btn_exit.pack(pady=10)

# Run the application
root.mainloop()
